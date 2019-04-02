# -*- coding:utf-8 -*-
# @Time:2019/2/24/002421:28
# @Author:tata
# @Email:286380647@qq.com
# @File:tata_homework6_doexcel.py
# @Software:PyCharm Community Edition

# 2：思考：分别将我们学过的数据类型  int float boolean str list tuple dict 写到每个单元格里面，
# 观察，你通过openpyxl操作后拿到的数据分别是是什么类型。

# 答:
#int float为对应的int float值
#boolean str list tuple dict 拿到的都为字符串



#第一题
from openpyxl import load_workbook
from openpyxl import Workbook
from API_1.common.configparser_ import ConF
from API_1.common.logging_ import MyLog
from API_1.common import project_path

class DoExcel:
    """完成Excel数据的读写 新建表单的操作"""
    def __init__(self,file_name,sheet_name):
        self.wb = load_workbook(file_name)
        self.sheet = self.wb[sheet_name]
    #函数一：读取指定表单的数据,按行读取
    def read_excel(self):

        all_row_list = []

        button = ConF(project_path.conf_path).read_str("case",'button')
        # MyLog().info("我要开始读取数据了")
        if button == '[all]':
            for item in range(2,self.sheet.max_row+1):
                row_list = []
                row_list.append(self.sheet.cell(item,1).value)   #用例的序号
                row_list.append(self.sheet.cell(item,2).value)  #用例标题
                row_list.append(self.sheet.cell(item,3).value)  #用例的模块
                row_list.append(self.sheet.cell(item,4).value)  #方法
                row_list.append(self.sheet.cell(item,5).value)  #url
                if self.sheet.cell(item,6).value.find("tel")!=-1 : #如果存在"tel",则将其替换为对应的手机号
                # if "tel" in self.sheet.cell(item, 6).value:
                    #将"tel"替换为表格里tel里的数据
                    row_list.append(self.sheet.cell(item,6).value.replace("tel",str(self.wb["tel"].cell(1,2).value)))
                    self.wb["tel"].cell(1, 2).value = self.wb["tel"].cell(1,2).value+1 #表格中tel每次使用后进行+1处理
                    self.wb.save(project_path.cases_path)
                    self.wb.close()
                else:
                    row_list.append(self.sheet.cell(item, 6).value)
                row_list.append(self.sheet.cell(item,7).value)  #sql
                row_list.append(self.sheet.cell(item,8).value)  #期望结果
                all_row_list.append(row_list)
        else:
            for item in eval(button):#
                row_list = []
                row_list.append(self.sheet.cell(item+1, 1).value)  # 用例的序号
                row_list.append(self.sheet.cell(item+1, 2).value)  # 用例标题
                row_list.append(self.sheet.cell(item+1, 3).value)  #用例的模块
                row_list.append(self.sheet.cell(item+1, 4).value)  # 方法
                row_list.append(self.sheet.cell(item+1, 5).value)  # url
                # if self.sheet.cell(item+1,6).value.find("tel")!=-1 : #如果存在"tel",则将其替换为对应的手机号
                if "tel" in sheet.cell(item, 6).value:
                    #将"tel"替换为表格里tel里的数据
                    row_list.append(self.sheet.cell(item+1,6).value.replace("tel",str(self.wb["tel"].cell(1,2).value)))
                    self.wb["tel"].cell(1, 2).value = self.wb["tel"].cell(1,2).value+1 #表格中tel每次使用后进行+1处理
                    self.wb.save(project_path.cases_path)
                    self.wb.close()
                else:
                    row_list.append(self.sheet.cell(item+1, 6).value)
                row_list.append(self.sheet.cell(item+1,7).value)  #sql
                row_list.append(self.sheet.cell(item+1,8).value)  #期望结果
                all_row_list.append(row_list)
        self.wb.close()
        # MyLog().info("读取数据了完毕")
        return all_row_list


    # 函数二：在指定的单元格写入指定的数据，并保存到当前Excel
    def write_excel(self,row,column,value):
        # MyLog().info("我要开始写入数据啦")
        try:
            self.sheet.cell(row,column).value = value
            self.wb.save(project_path.cases_path)
            self.wb.close()
            # MyLog().info("数据写入完毕")
        except Exception as e:
            MyLog().error("数据写入错误",e)


    # 函数三：新建一个Excel
    def new_file(self,file_name):
        # MyLog().info("我要开始新建一个excel表格")
        wb = Workbook()
        wb.save(file_name)
        wb.close()
        # MyLog().info("新建完毕")


if __name__ == '__main__':
    excel = DoExcel(project_path.cases_path,"test")
    excel = DoExcel(project_path.cases_path,"投资")
    excel = DoExcel(project_path.cases_path,"新增项目")
    # # excel.new_file("11.xlsx")          #新建一个 11.xlsx 的表格
    # excel.write_excel(2,9,"这是A1")  #修改指定表格
    print(excel.read_excel())  #按行读取数据,可在配置文件log.conf里的button修改读取用例的条件

